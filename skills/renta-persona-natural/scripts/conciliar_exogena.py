#!/usr/bin/env python3
"""
conciliar_exogena.py
Motor de conciliación automática y cruce de información entre los extractos/certificados
del contribuyente (transacciones_depuradas.csv) y la Información Exógena de la DIAN.
"""

import sys
import csv
import json
import argparse
import re
from pathlib import Path
from typing import Dict, List, Any, Tuple


def normalize_nit(nit_str: Any) -> str:
    if not nit_str:
        return ""
    # Extraer sólo los dígitos antes del guión o completos
    raw = str(nit_str).split("-")[0].strip()
    return re.sub(r"\D", "", raw)


def parse_exogena_file(exogena_path: str) -> List[Dict[str, Any]]:
    path = Path(exogena_path)
    if not path.exists():
        raise FileNotFoundError(f"No se encontró el archivo de exógena en: {exogena_path}")

    records = []
    if path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        # Buscar fila de encabezado
        start_row = 1
        for r in range(1, min(25, ws.max_row + 1)):
            v1 = str(ws.cell(row=r, column=1).value or "").lower()
            v2 = str(ws.cell(row=r, column=2).value or "").lower()
            if "nit" in v1 or "persona que reporta" in v1 or "nombre" in v2:
                start_row = r + 1
                break

        for r in range(start_row, ws.max_row + 1):
            nit_reporta = ws.cell(row=r, column=1).value
            nombre_reporta = ws.cell(row=r, column=2).value
            detalle = ws.cell(row=r, column=5).value or ws.cell(row=r, column=4).value or ""
            val_raw = ws.cell(row=r, column=6).value
            info_adic = ws.cell(row=r, column=8).value or ""

            if not nit_reporta and not nombre_reporta:
                continue

            try:
                val = float(str(val_raw).replace(",", "").replace("$", "").strip()) if val_raw is not None else 0.0
            except ValueError:
                val = 0.0

            records.append({
                "nit": normalize_nit(nit_reporta),
                "nit_raw": str(nit_reporta or ""),
                "nombre_tercero": str(nombre_reporta or "").strip(),
                "detalle_concepto": str(detalle).strip(),
                "valor_cop": val,
                "info_adicional": str(info_adic).strip()
            })

    elif path.suffix.lower() == ".csv":
        with open(path, mode="r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                nit_rep = row.get("NIT", row.get("nit", ""))
                nombre_rep = row.get("Nombre", row.get("tercero_nombre", ""))
                detalle = row.get("Detalle", row.get("concepto", ""))
                val_raw = row.get("Valor", row.get("valor_cop", "0"))
                try:
                    val = float(str(val_raw).replace(",", "").replace("$", "").strip())
                except ValueError:
                    val = 0.0

                records.append({
                    "nit": normalize_nit(nit_rep),
                    "nit_raw": str(nit_rep),
                    "nombre_tercero": str(nombre_rep).strip(),
                    "detalle_concepto": str(detalle).strip(),
                    "valor_cop": val,
                    "info_adicional": row.get("Informacion_Adicional", "")
                })

    return records


def parse_facturas_electronicas(facturas_path: str) -> Tuple[float, float, int]:
    path = Path(facturas_path)
    if not path.exists():
        return 0.0, 0.0, 0

    total_facturado = 0.0
    total_susceptible = 0.0
    count_elec = 0

    if path.suffix.lower() in (".xlsx", ".xlsm"):
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        # Buscar encabezado
        start_row = 26
        for r in range(15, min(35, ws.max_row + 1)):
            c1 = str(ws.cell(row=r, column=1).value or "").lower()
            if "emisor" in c1 or "identificaci" in c1:
                start_row = r + 1
                break

        for r in range(start_row, ws.max_row + 1):
            val_fact = ws.cell(row=r, column=4).value or 0
            val_susc = ws.cell(row=r, column=8).value or 0
            medio = str(ws.cell(row=r, column=9).value or "")

            try:
                vf = float(val_fact)
                vs = float(val_susc)
            except ValueError:
                continue

            total_facturado += vf
            total_susceptible += vs
            if "electr" in medio.lower():
                count_elec += 1

    return total_facturado, total_susceptible, count_elec


def conciliar_transacciones_con_exogena(csv_transacciones: str, exogena_path: str,
                                        facturas_path: str = None) -> Dict[str, Any]:
    # 1. Cargar transacciones del usuario
    trans_path = Path(csv_transacciones)
    if not trans_path.exists():
        raise FileNotFoundError(f"No existe el archivo de transacciones en: {csv_transacciones}")

    transacciones = []
    with open(trans_path, mode="r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for r in reader:
            transacciones.append(dict(r))

    # 2. Cargar registros de exógena
    exogena_records = parse_exogena_file(exogena_path)

    # 3. Conciliar
    # Mapear transacciones por NIT normalizado
    matched_exogena_indices = set()
    conciliacion_items = []
    
    for t in transacciones:
        t_nit = normalize_nit(t.get("tercero_nit", ""))
        raw_val = t.get("valor_cop", "0").replace("$", "").replace("'", "").replace(".", "").replace(",", "").strip()
        try:
            t_val = float(raw_val)
        except ValueError:
            t_val = 0.0

        # Buscar coincidencia en exógena
        best_match_idx = None
        best_diff = float("inf")
        
        for idx, exo in enumerate(exogena_records):
            if idx in matched_exogena_indices:
                continue
            
            exo_nit = exo["nit"]
            if t_nit and exo_nit and (t_nit == exo_nit or t_nit.startswith(exo_nit) or exo_nit.startswith(t_nit)):
                diff = abs(t_val - exo["valor_cop"])
                if diff < best_diff:
                    best_diff = diff
                    best_match_idx = idx

        if best_match_idx is not None and best_diff <= 5000:
            exo_match = exogena_records[best_match_idx]
            matched_exogena_indices.add(best_match_idx)
            estado = "MATCH_EXACTO" if best_diff <= 1000 else "DIFERENCIA_VALOR"
            
            t["estado_exogena"] = estado
            t["valor_exogena_cop"] = exo_match["valor_cop"]
            t["diferencia_exogena_cop"] = t_val - exo_match["valor_cop"]
            t["resolucion_usuario"] = "CONFIRMADO_CERTIFICADO"

            conciliacion_items.append({
                "id": f"CONC-{len(conciliacion_items)+1:03d}",
                "tercero_nit": t.get("tercero_nit", ""),
                "tercero_nombre": t.get("tercero_nombre", exo_match["nombre_tercero"]),
                "concepto": t.get("descripcion", exo_match["detalle_concepto"]),
                "valor_certificado": t_val,
                "valor_exogena": exo_match["valor_cop"],
                "diferencia": t_val - exo_match["valor_cop"],
                "estado": estado,
                "resolucion_usuario": "CONFIRMADO_CERTIFICADO",
                "observaciones": "Coincidencia automática validada con DIAN"
            })
        else:
            t["estado_exogena"] = "SOLO_EN_CERTIFICADOS"
            t["valor_exogena_cop"] = 0
            t["diferencia_exogena_cop"] = t_val
            t["resolucion_usuario"] = "CONFIRMADO_CERTIFICADO"
            
            conciliacion_items.append({
                "id": f"CONC-{len(conciliacion_items)+1:03d}",
                "tercero_nit": t.get("tercero_nit", ""),
                "tercero_nombre": t.get("tercero_nombre", ""),
                "concepto": t.get("descripcion", ""),
                "valor_certificado": t_val,
                "valor_exogena": 0.0,
                "diferencia": t_val,
                "estado": "SOLO_EN_CERTIFICADOS",
                "resolucion_usuario": "CONFIRMADO_CERTIFICADO",
                "observaciones": "Partida presente en certificados/extractos pero no encontrada en reporte de exógena"
            })

    # 4. Procesar partidas de exógena no emparejadas (SOLO_EN_EXOGENA)
    discrepancias_para_usuario = []
    for idx, exo in enumerate(exogena_records):
        if idx not in matched_exogena_indices:
            disc_id = f"DISC-{len(discrepancias_para_usuario)+1:02d}"
            item = {
                "id": f"CONC-{len(conciliacion_items)+1:03d}",
                "tercero_nit": exo["nit_raw"],
                "tercero_nombre": exo["nombre_tercero"],
                "concepto": exo["detalle_concepto"],
                "valor_certificado": 0.0,
                "valor_exogena": exo["valor_cop"],
                "diferencia": -exo["valor_cop"],
                "estado": "SOLO_EN_EXOGENA",
                "resolucion_usuario": "PENDIENTE_CONSULTA_USUARIO",
                "observaciones": exo["info_adicional"]
            }
            conciliacion_items.append(item)
            
            # Formular pregunta clara para el usuario
            pregunta = f"En la Exógena DIAN aparece reportado por {exo['nombre_tercero']} (NIT: {exo['nit_raw']}) el concepto '{exo['detalle_concepto']}' por ${exo['valor_cop']:,.0f} COP. ¿Deseas incluirlo en tu declaración?"
            discrepancias_para_usuario.append({
                "discrepancia_id": disc_id,
                "tercero": exo["nombre_tercero"],
                "nit": exo["nit_raw"],
                "concepto": exo["detalle_concepto"],
                "valor_exogena": exo["valor_cop"],
                "pregunta": pregunta,
                "sugerencia_cedula": "PATRIMONIO_BIENES" if "vehículo" in exo["detalle_concepto"].lower() or "inversión" in exo["detalle_concepto"].lower() else "RENTAS_CAPITAL"
            })

    # 5. Procesar facturas electrónicas si existen
    tot_facturado, tot_susceptible, num_facturas = (0.0, 0.0, 0)
    has_facturas = False
    if facturas_path and Path(facturas_path).exists():
        tot_facturado, tot_susceptible, num_facturas = parse_facturas_electronicas(facturas_path)
        has_facturas = True

    # 6. Escribir transacciones actualizadas con columnas de exógena
    fieldnames = list(transacciones[0].keys())
    for col in ["estado_exogena", "valor_exogena_cop", "diferencia_exogena_cop", "resolucion_usuario"]:
        if col not in fieldnames:
            fieldnames.append(col)

    with open(trans_path, mode="w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in transacciones:
            writer.writerow(r)

    # 7. Resumen y métricas
    total_matches = sum(1 for item in conciliacion_items if item["estado"] == "MATCH_EXACTO")
    total_diffs = sum(1 for item in conciliacion_items if item["estado"] == "DIFERENCIA_VALOR")
    total_solo_exo = sum(1 for item in conciliacion_items if item["estado"] == "SOLO_EN_EXOGENA")
    total_solo_cert = sum(1 for item in conciliacion_items if item["estado"] == "SOLO_EN_CERTIFICADOS")
    pct_match = (total_matches / len(exogena_records) * 100.0) if exogena_records else 0.0

    estado_resumen = {
        "has_exogena": True,
        "has_facturas_electronicas": has_facturas,
        "archivo_exogena": str(Path(exogena_path).name),
        "archivo_facturas": str(Path(facturas_path).name) if has_facturas else None,
        "total_susceptible_factura_elec": tot_susceptible,
        "deduccion_1pct_factura_elec": round(tot_susceptible * 0.01),
        "total_partidas_exogena": len(exogena_records),
        "total_conciliadas": total_matches,
        "total_discrepancias": total_diffs + total_solo_exo,
        "porcentaje_match": round(pct_match, 2),
        "metricas": {
            "match_exacto": total_matches,
            "diferencia_valor": total_diffs,
            "solo_en_exogena": total_solo_exo,
            "solo_en_certificados": total_solo_cert
        },
        "discrepancias_para_usuario": discrepancias_para_usuario,
        "items": conciliacion_items
    }

    return estado_resumen


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Concilia transacciones con Información Exógena DIAN")
    parser.add_argument("csv_file", help="Ruta a transacciones_depuradas.csv")
    parser.add_argument("exogena_file", help="Ruta a DIAN - Informacion exogena.xlsx o .csv")
    parser.add_argument("--facturas", default=None, help="Ruta a DIAN - Facturas electronicas.xlsx")
    parser.add_argument("--out-csv", default="conciliacion_exogena.csv", help="Ruta para exportar libro de conciliación CSV")
    parser.add_argument("--out-json", default="estado_conciliacion.json", help="Ruta para exportar snapshot JSON")

    args = parser.parse_args()

    resultado = conciliar_transacciones_con_exogena(
        csv_transacciones=args.csv_file,
        exogena_path=args.exogena_file,
        facturas_path=args.facturas
    )

    # Exportar JSON
    with open(args.out_json, "w", encoding="utf-8") as f:
        json.dump(resultado, f, indent=2, ensure_ascii=False)

    # Exportar CSV de auditoría
    if resultado.get("items"):
        with open(args.out_csv, "w", encoding="utf-8", newline="") as f:
            fieldnames = ["id", "tercero_nit", "tercero_nombre", "concepto", "valor_certificado", "valor_exogena", "diferencia", "estado", "resolucion_usuario", "observaciones"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for it in resultado["items"]:
                writer.writerow(it)

    print("\n=======================================================")
    print("  [OK] CONCILIACIÓN CON INFORMACIÓN EXÓGENA DIAN")
    print("=======================================================")
    print(f" Total Partidas Exógena:   {resultado['total_partidas_exogena']}")
    print(f" Coincidencias Exactas:    {resultado['metricas']['match_exacto']} ({resultado['porcentaje_match']}%)")
    print(f" Diferencias de Valor:     {resultado['metricas']['diferencia_valor']}")
    print(f" Sólo en Exógena:          {resultado['metricas']['solo_en_exogena']}")
    print(f" Sólo en Certificados:     {resultado['metricas']['solo_en_certificados']}")
    if resultado['has_facturas_electronicas']:
        print(f" Compras Factura Electr:   ${resultado['total_susceptible_factura_elec']:,.0f} COP (Deducción 1%: ${resultado['deduccion_1pct_factura_elec']:,.0f} COP)")
    print("-------------------------------------------------------")
    print(f" Reporte CSV:              {args.out_csv}")
    print(f" Snapshot JSON:            {args.out_json}")
    print("=======================================================\n")
